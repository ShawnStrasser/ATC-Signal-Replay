#!/usr/bin/env python
"""
firmware_validate.py — Standalone firmware validation script.

Replays event logs to controllers with new firmware, compares output to
original logs, and generates an HTML report with divergence charts.

Usage:
    python firmware_validate.py                  # interactive, uses settings.json
    python firmware_validate.py --verbose        # extra debug output
    python firmware_validate.py --report-only    # skip replay, just run analysis
    python firmware_validate.py --settings custom_settings.json

Settings are loaded from settings.json (editable JSON file in the same folder).
"""

from __future__ import annotations

import argparse
import json
import sys
import time
import traceback
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import duckdb
import pandas as pd
import requests
from openpyxl import load_workbook

import signal_replay as sr
from signal_replay.ntcip import send_ntcip
from signal_replay.report import generate_report

# ---------------------------------------------------------------------------
# Dev / test mode — limit to these devices for fast iteration
# ---------------------------------------------------------------------------
DEV_DEVICES = ["2B085", "08042", "2C042"]

# ---------------------------------------------------------------------------
# Logging helpers
# ---------------------------------------------------------------------------
_VERBOSE = False


def log(msg: str, *, always: bool = True) -> None:
    """Print a message. If always=False, only prints in verbose mode."""
    if always or _VERBOSE:
        print(msg, flush=True)


def vlog(msg: str) -> None:
    """Verbose-only log."""
    log(msg, always=False)


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------
def load_settings(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Catalog helpers
# ---------------------------------------------------------------------------
def read_catalog(catalog_path: Path) -> List[dict]:
    wb = load_workbook(catalog_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [str(h).strip().lower() for h in rows[0]]
    catalog: List[dict] = []
    for row in rows[1:]:
        if not any(row):
            continue
        r = dict(zip(header, [str(v).strip() if v is not None else "" for v in row]))
        raw_cl = r.get("cyclelength", "")
        raw_off = r.get("offset", "")
        cycle_length = int(float(raw_cl)) if raw_cl not in ("", "None") else 0
        offset = float(raw_off) if raw_off not in ("", "None") else 0.0
        catalog.append({
            "TSSU": r.get("tssu", ""),
            "Version": r.get("version", ""),
            "Type": r.get("type", "Similarity") or "Similarity",
            "CycleLength": cycle_length,
            "Offset": offset,
            "Notes": r.get("notes", ""),
        })
    return catalog


def find_log(tssu: str, logs_dir: Path) -> Optional[Path]:
    for ext in (".parquet", ".csv", ".db"):
        p = logs_dir / f"{tssu}{ext}"
        if p.exists():
            return p
    matches = sorted(logs_dir.glob(f"{tssu}*"))
    return matches[0] if matches else None


def find_database(tssu: str, databases_dir: Path) -> Optional[Path]:
    matches = sorted(databases_dir.glob(f"{tssu}*"))
    return matches[0] if matches else None


def normalize_target(target: str) -> str:
    parts = target.split(":")
    if len(parts) == 2:
        host, port_text = parts
        port = int(port_text)
        return f"{host}:{port}:{port}"
    if len(parts) == 3:
        host, udp_text, http_text = parts
        return f"{host}:{int(udp_text)}:{int(http_text)}"
    raise ValueError(f"Invalid controller target: {target}")


# ---------------------------------------------------------------------------
# Build suite
# ---------------------------------------------------------------------------
def build_suite(
    settings: dict,
    firmware_dir: Path,
    catalog: List[dict],
    conflict_pairs: dict,
) -> Tuple[sr.FirmwareTestSuite, List[sr.TestScenario], List[sr.TestBatch], dict]:
    """Build scenarios, batches, and the full test suite. Returns (suite, scenarios, batches, file_map)."""

    logs_dir = firmware_dir / settings["logs_dir"]
    databases_dir = firmware_dir / settings["databases_dir"]
    controller_targets = settings["controller_targets"]
    firmware_version = settings["firmware_version"]
    baseline_version = settings["baseline_version"]
    batch_size = len(controller_targets)

    file_map: Dict[str, dict] = {}
    for r in catalog:
        tssu = r["TSSU"]
        file_map[tssu] = {
            "log": find_log(tssu, logs_dir),
            "db": find_database(tssu, databases_dir),
        }

    scenarios: List[sr.TestScenario] = []
    skipped: List[str] = []

    for r in catalog:
        tssu = r["TSSU"]
        log_path = file_map[tssu]["log"]
        db_path = file_map[tssu]["db"]
        if not log_path:
            skipped.append(tssu)
            continue

        test_type = sr.TestType.CONFLICT if r["Type"].strip().lower() == "conflict" else sr.TestType.SIMILARITY
        cycle_length = r.get("CycleLength", 0) or 0
        offset = r.get("Offset", 0.0) or 0.0
        tod_align = cycle_length == 0

        kwargs: dict = {}
        if test_type == sr.TestType.CONFLICT:
            kwargs["replays"] = 25
            pairs_key = tssu if tssu in conflict_pairs else tssu.rstrip("_c") if tssu.endswith("_c") else tssu
            if pairs_key in conflict_pairs:
                kwargs["incompatible_pairs"] = conflict_pairs[pairs_key]

        scenarios.append(sr.TestScenario(
            scenario_id=tssu,
            database_name=str(db_path) if db_path else f"{tssu}.bin",
            events_source=str(log_path),
            test_type=test_type,
            description=f"{r['Type']} | {r['Notes']}" if r["Notes"] else r["Type"],
            notes_column=r["Notes"],
            cycle_length=cycle_length,
            cycle_offset=offset,
            tod_align=tod_align,
            **kwargs,
        ))

    if skipped:
        log(f"Skipped (no log): {', '.join(skipped)}")

    normalized_targets = [normalize_target(t) for t in controller_targets]
    batches: List[sr.TestBatch] = []
    for i in range(0, len(scenarios), batch_size):
        chunk = scenarios[i:i + batch_size]
        batch_num = i // batch_size + 1
        targets = normalized_targets[:len(chunk)]
        batches.append(sr.TestBatch(
            batch_id=f"batch_{batch_num}",
            assignments={s.scenario_id: t for s, t in zip(chunk, targets)},
            description=f"Batch {batch_num}: {len(chunk)} scenarios",
        ))

    comp = settings.get("comparison", {})
    suite = sr.FirmwareTestSuite(
        suite_name="Firmware Validation",
        firmware_version=firmware_version,
        baseline_version=baseline_version,
        scenarios=scenarios,
        batches=batches,
        output_dir=str(firmware_dir / settings["results_dir"]),
        comparison_thresholds=sr.ComparisonThresholds(
            sequence_threshold=comp.get("sequence_threshold", 0.05),
            timing_threshold=comp.get("timing_threshold", 0.02),
            match_threshold=comp.get("match_threshold", 95.0),
        ),
    )

    return suite, scenarios, batches, file_map


# ---------------------------------------------------------------------------
# Controller check
# ---------------------------------------------------------------------------
def parse_target(target: str) -> Tuple[str, int, int]:
    parts = target.split(":")
    if len(parts) == 2:
        ip, port_text = parts
        port = int(port_text)
        return ip, port, port
    if len(parts) == 3:
        ip, udp_text, http_text = parts
        return ip, int(udp_text), int(http_text)
    raise ValueError(f"Invalid target format: {target}")


def check_controller(target: str) -> bool:
    ip, udp_port, http_port = parse_target(target)
    try:
        send_ntcip((ip, udp_port), 1, 1, "Vehicle", timeout=1.0)
        send_ntcip((ip, udp_port), 1, 0, "Vehicle", timeout=1.0)
        snmp_ok = True
    except Exception:
        snmp_ok = False
    try:
        r = requests.get(f"http://{ip}:{http_port}/v1/asclog/xml/full", timeout=2.0, verify=False)
        http_ok = r.status_code == 200
    except Exception:
        http_ok = False
    return snmp_ok and http_ok


def wait_for_controllers(targets: List[str], labels: List[str]) -> None:
    """Poll controllers until all respond OK. Print status each cycle."""
    log("Polling controllers...")
    while True:
        with ThreadPoolExecutor() as ex:
            statuses = list(ex.map(check_controller, targets))
        parts = []
        for ok, lbl in zip(statuses, labels):
            parts.append(f"  OK   {lbl}" if ok else f"  FAIL {lbl}")
        print("\r\033[K" + "\n".join(parts), flush=True)
        if all(statuses):
            log("All controllers responding.")
            break
        time.sleep(3)
        # Move cursor up to overwrite
        print(f"\033[{len(parts)}A", end="", flush=True)


# ---------------------------------------------------------------------------
# Batch operations
# ---------------------------------------------------------------------------
def detect_current_batch(suite: sr.FirmwareTestSuite, batches: List[sr.TestBatch]) -> Tuple[Optional[sr.TestBatch], set]:
    checkpoint_path = Path(suite.output_dir) / suite.firmware_version / "checkpoint.json"
    completed: set = set()
    if checkpoint_path.exists():
        with open(checkpoint_path) as f:
            ck = json.load(f)
        completed = set(ck.get("completed_batches", []))

    for b in batches:
        if b.batch_id not in completed:
            return b, completed
    return None, completed


def run_batch(suite: sr.FirmwareTestSuite, batch: sr.TestBatch) -> dict:
    runner = sr.BatchRunner(suite, debug=_VERBOSE)

    def auto_db_loader(db_name: str, target: str) -> bool:
        vlog(f"  [auto] DB {db_name} -> {target}")
        return True

    return runner.run(
        db_loader_callback=auto_db_loader,
        batch_ids=[batch.batch_id],
    )


# ---------------------------------------------------------------------------
# Analysis (comparison) — designed for multiprocessing
# ---------------------------------------------------------------------------
def _extract_collected_events(
    suite: sr.FirmwareTestSuite,
    cp: dict,
    tmp_dir: Path,
) -> Dict[str, Path]:
    """
    Pre-extract collected events from DuckDB into temporary parquet files.
    This avoids DuckDB file-locking issues when using multiprocessing on Windows.
    Returns a mapping of scenario_id -> parquet path.
    """
    tmp_dir.mkdir(parents=True, exist_ok=True)
    # Group scenarios by their duckdb file to minimize open/close cycles
    db_to_scenarios: Dict[str, List[str]] = {}
    for scenario in suite.scenarios:
        db_path = cp["scenario_db_map"].get(scenario.scenario_id)
        if db_path:
            db_to_scenarios.setdefault(db_path, []).append(scenario.scenario_id)

    extracted: Dict[str, Path] = {}
    for db_path, scenario_ids in db_to_scenarios.items():
        vlog(f"  Extracting from {Path(db_path).name}: {len(scenario_ids)} scenarios")
        con = duckdb.connect(db_path, read_only=True)
        try:
            for sid in scenario_ids:
                df = con.execute(
                    "SELECT * FROM events WHERE device_id = ? ORDER BY timestamp",
                    [sid],
                ).df()
                out_path = tmp_dir / f"{sid}_collected.parquet"
                df.to_parquet(out_path, index=False)
                extracted[sid] = out_path
        finally:
            con.close()
    return extracted


def _compare_one_scenario(args: Tuple) -> dict:
    """
    Worker function for ProcessPoolExecutor.
    All arguments are plain strings/numbers to avoid pickle issues.
    Reads from parquet files (no DuckDB needed in workers).
    """
    (scenario_id, events_source, test_type_str, collected_parquet,
     firmware_version, plots_dir_str, settle_minutes,
     max_plots, window_minutes, verbose, notes_column) = args

    import signal_replay as sr
    import pandas as pd
    import os, contextlib

    test_type = sr.TestType.CONFLICT if test_type_str == "CONFLICT" else sr.TestType.SIMILARITY

    original = sr.load_events(events_source)
    collected = pd.read_parquet(collected_parquet)

    result = sr.compare_runs(
        events_a=original, events_b=collected,
        device_id=scenario_id,
        run_a_label="original_logs", run_b_label=firmware_version,
        auto_align=True, settle_minutes=settle_minutes,
    )

    plot_paths: list = []
    timeline_a = timeline_b = None
    if not original.empty and not collected.empty:
        try:
            # Suppress atspm's verbose stdout unless --verbose
            _devnull = open(os.devnull, "w") if not verbose else None
            _ctx = contextlib.redirect_stdout(_devnull) if _devnull else contextlib.nullcontext()
            with _ctx:
                timeline_a = sr.generate_timeline(original, device_id=scenario_id)
                timeline_b = sr.generate_timeline(collected, device_id=scenario_id)
            if _devnull:
                _devnull.close()
            remove_events = [
                "Ped Omit", "Phase Hold", "Phase Omit", "Phase Call",
                "Transition Longway", "Transition Shortway",
            ]
            timeline_a = timeline_a[~timeline_a["EventClass"].isin(remove_events)]
            timeline_b = timeline_b[~timeline_b["EventClass"].isin(remove_events)]
        except Exception as e:
            if verbose:
                print(f"    Timeline generation failed for {scenario_id}: {e}", flush=True)

    if result.divergence_windows and timeline_a is not None and not timeline_a.empty and not timeline_b.empty:
        try:
            _devnull2 = open(os.devnull, "w") if not verbose else None
            _ctx2 = contextlib.redirect_stdout(_devnull2) if _devnull2 else contextlib.nullcontext()
            with _ctx2:
                time_offset_b = sr.compute_timeline_offset(timeline_a, timeline_b)
                plot_paths = sr.create_multi_divergence_plots(
                    timeline_a=timeline_a, timeline_b=timeline_b,
                    comparison_result=result, output_dir=plots_dir_str,
                    label_a="Original", label_b=firmware_version,
                    max_plots=max_plots, window_minutes=window_minutes,
                    time_offset_b=time_offset_b,
                )
            if _devnull2:
                _devnull2.close()
        except Exception as e:
            if verbose:
                print(f"    Chart generation failed for {scenario_id}: {e}", flush=True)
                import traceback as _tb
                _tb.print_exc()

    passed = result.match_percentage >= 95.0
    phase_diffs: list = []
    if not passed and timeline_a is not None and not timeline_a.empty and not timeline_b.empty:
        try:
            # The two timelines come from different absolute dates (original vs.
            # collected replay). We need to align them by relative time from
            # their respective starts before settle-trimming and overlap-clipping.
            start_a = timeline_a["StartTime"].min()
            start_b = timeline_b["StartTime"].min()

            tl_a_rel = timeline_a.copy()
            tl_b_rel = timeline_b.copy()

            # Convert to a common epoch (use start_a as the reference)
            tl_b_rel["StartTime"] = start_a + (tl_b_rel["StartTime"] - start_b)
            tl_b_rel["EndTime"] = start_a + (tl_b_rel["EndTime"] - start_b)

            settle_td = pd.Timedelta(minutes=settle_minutes)
            tl_a_settled = tl_a_rel[tl_a_rel["StartTime"] >= start_a + settle_td].copy()
            tl_b_settled = tl_b_rel[tl_b_rel["StartTime"] >= start_a + settle_td].copy()

            overlap_end = min(tl_a_settled["EndTime"].max(), tl_b_settled["EndTime"].max())
            tl_a_settled = tl_a_settled[tl_a_settled["StartTime"] <= overlap_end]
            tl_b_settled = tl_b_settled[tl_b_settled["StartTime"] <= overlap_end]

            if verbose:
                signal_classes = {'Green', 'Yellow', 'Red', 'Overlap Green', 'Overlap Trail Green', 'Overlap Yellow', 'Overlap Red'}
                sig_a = tl_a_settled[tl_a_settled["EventClass"].isin(signal_classes)]
                sig_b = tl_b_settled[tl_b_settled["EventClass"].isin(signal_classes)]
                print(f"    [diag] timeline_a range: {timeline_a['StartTime'].min()} to {timeline_a['EndTime'].max()}", flush=True)
                print(f"    [diag] timeline_b range: {timeline_b['StartTime'].min()} to {timeline_b['EndTime'].max()}", flush=True)
                print(f"    [diag] after settle+overlap: tl_a={len(tl_a_settled)} ({len(sig_a)} signal), "
                      f"tl_b={len(tl_b_settled)} ({len(sig_b)} signal), overlap_end={overlap_end}", flush=True)

            phase_diffs = sr.generate_phase_difference_summary(tl_a_settled, tl_b_settled, tolerance_seconds=0.2)
        except Exception as e:
            if verbose:
                print(f"    Phase breakdown failed for {scenario_id}: {e}", flush=True)
                import traceback as _tb
                _tb.print_exc()

    # Truncate the summary to show at most 10 divergences
    raw_summary = result.format_summary()
    summary_lines = raw_summary.split("\n")
    # Find where divergence list starts (indented lines after "Divergences: N")
    truncated_lines = []
    div_count = 0
    max_div_shown = 5
    for line in summary_lines:
        if line.startswith("  ") and div_count >= max_div_shown:
            continue  # skip excess divergence lines
        truncated_lines.append(line)
        if line.startswith("  "):
            div_count += 1
            if div_count == max_div_shown and len(result.divergence_windows) > max_div_shown:
                truncated_lines.append(f"  ... and {len(result.divergence_windows) - max_div_shown} more divergences")

    return {
        "scenario_id": scenario_id,
        "test_type_str": test_type_str,
        "passed": passed,
        "match_percentage": result.match_percentage,
        "num_divergences": len(result.divergence_windows),
        "summary": "\n".join(truncated_lines),
        "plot_paths": plot_paths,
        "phase_diffs": phase_diffs,
        "notes_column": notes_column,
        "chunk_scores": [
            {"center_seconds": c.center_seconds,
             "match_percentage": c.match_percentage,
             "window_seconds": c.window_seconds}
            for c in result.chunk_scores
        ],
        "temporal_shift_seconds": result.temporal_shift_seconds,
        "sparkline_svg": sr.render_sparkline_svg(
            result.chunk_scores,
            pass_threshold=95.0,
        ) if result.chunk_scores else "",
    }


def run_analysis(
    suite: sr.FirmwareTestSuite,
    settings: dict,
    firmware_dir: Path,
) -> List[sr.ScenarioResult]:
    """Run comparisons with multiprocessing. Returns list of ScenarioResult."""
    import tempfile
    import shutil

    firmware_version = suite.firmware_version
    comp = settings.get("comparison", {})
    settle_minutes = comp.get("settle_minutes", 3.0)
    max_plots = comp.get("max_divergence_plots", 3)
    window_minutes = comp.get("divergence_window_minutes", 5.0)
    max_workers = settings.get("analysis_workers", 4)

    # Load checkpoint
    checkpoint_path = Path(suite.output_dir) / firmware_version / "checkpoint.json"
    with open(checkpoint_path) as f:
        cp = json.load(f)

    completed = cp.get("completed_batches", [])
    total_batches = [b.batch_id for b in suite.batches]
    if set(total_batches) - set(completed):
        missing = sorted(set(total_batches) - set(completed))
        log(f"WARNING: Not all batches complete. Missing: {missing}")
        log("         Report will only cover completed batches.")

    plots_dir = Path(suite.output_dir) / firmware_version / "divergence_plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    # Pre-extract collected events from DuckDB to avoid file-locking in workers
    tmp_dir = Path(suite.output_dir) / firmware_version / "_tmp_analysis"
    log("Extracting collected events from DuckDB...")
    extracted_map = _extract_collected_events(suite, cp, tmp_dir)
    log(f"Extracted {len(extracted_map)} scenarios to temp parquet files.")

    # Build job list — all plain types for pickling (no DuckDB paths)
    jobs: list = []
    for scenario in suite.scenarios:
        collected_parquet = extracted_map.get(scenario.scenario_id)
        if not collected_parquet:
            log(f"  No output for {scenario.scenario_id}, skipping")
            continue
        test_type_str = "CONFLICT" if scenario.test_type == sr.TestType.CONFLICT else "SIMILARITY"
        jobs.append((
            scenario.scenario_id,
            scenario.events_source,
            test_type_str,
            str(collected_parquet),
            firmware_version,
            str(plots_dir),
            settle_minutes,
            max_plots,
            window_minutes,
            _VERBOSE,
            scenario.notes_column,
        ))

    if not jobs:
        log("No scenarios to analyze.")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return []

    n_workers = min(max_workers, len(jobs))
    log(f"\nAnalyzing {len(jobs)} scenarios with {n_workers} workers...")
    sys.stdout.flush()

    results: List[sr.ScenarioResult] = []
    done = 0

    try:
        with ProcessPoolExecutor(max_workers=n_workers) as executor:
            futures = {
                executor.submit(_compare_one_scenario, job): job[0]
                for job in jobs
            }
            for future in as_completed(futures):
                sid = futures[future]
                done += 1
                try:
                    out = future.result()
                    test_type = sr.TestType.CONFLICT if out["test_type_str"] == "CONFLICT" else sr.TestType.SIMILARITY
                    results.append(sr.ScenarioResult(
                        scenario_id=out["scenario_id"],
                        test_type=test_type,
                        firmware_version=firmware_version,
                        passed=out["passed"],
                        match_percentage=out["match_percentage"],
                        num_divergences=out["num_divergences"],
                        notes=out["summary"],
                        notes_column=out.get("notes_column", ""),
                        plot_paths=out["plot_paths"],
                        phase_differences=out["phase_diffs"],
                        runs_completed=1,
                        total_runs=1,
                        chunk_scores=out.get("chunk_scores", []),
                        sparkline_svg=out.get("sparkline_svg", ""),
                        temporal_shift_seconds=out.get("temporal_shift_seconds", 0.0),
                    ))
                    status = "PASS" if out["passed"] else "FAIL"
                    plots_msg = f", {len(out['plot_paths'])} charts" if out["plot_paths"] else ""
                    diffs_msg = ""
                    if out["phase_diffs"]:
                        diffs_msg = f"\n    Phase/overlap differences ({len(out['phase_diffs'])} phases):\n"
                        diffs_msg += sr.format_phase_differences(
                            out["phase_diffs"], label_a="Original", label_b=firmware_version
                        )
                    log(f"  [{done}/{len(jobs)}] {out['scenario_id']}: {out['match_percentage']:.1f}%  {status}  ({out['num_divergences']} divergences{plots_msg}){diffs_msg}")
                except Exception as e:
                    log(f"  [{done}/{len(jobs)}] {sid}: ERROR — {e}")
                    if _VERBOSE:
                        traceback.print_exc()
    finally:
        # Clean up temp parquet files
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)
        vlog("Cleaned up temp analysis files.")

    # Sort results by scenario name for consistent ordering in report
    results.sort(key=lambda r: r.scenario_id)
    return results


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------
def build_report(
    results: List[sr.ScenarioResult],
    suite: sr.FirmwareTestSuite,
) -> Path:
    report_dir = Path(suite.output_dir) / suite.firmware_version
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / "report.html"
    generate_report(results, suite, str(report_path))
    return report_path


# ---------------------------------------------------------------------------
# Archive + extract baseline
# ---------------------------------------------------------------------------
def archive_and_extract(suite: sr.FirmwareTestSuite, firmware_dir: Path) -> None:
    fw_ver = suite.firmware_version
    logs_dir = firmware_dir / "logs"
    archive_dir = firmware_dir / f"logs_{fw_ver}"

    if archive_dir.exists():
        log(f"Archive folder already exists: {archive_dir}")
        log("Skipping archive step.")
    elif logs_dir.exists():
        logs_dir.rename(archive_dir)
        log(f"Archived logs/ -> {archive_dir}")
    else:
        log("No logs/ folder found — nothing to archive.")

    logs_dir.mkdir(parents=True, exist_ok=True)
    log("Fresh logs/ directory ready.")

    checkpoint_path = Path(suite.output_dir) / fw_ver / "checkpoint.json"
    with open(checkpoint_path) as f:
        cp = json.load(f)

    extracted = 0
    for scenario in suite.scenarios:
        db_path = cp["scenario_db_map"].get(scenario.scenario_id)
        if not db_path:
            continue
        con = duckdb.connect(db_path)
        try:
            df = con.execute(
                "SELECT * FROM events WHERE device_id = ? ORDER BY timestamp",
                [scenario.scenario_id],
            ).df()
        finally:
            con.close()
        if df.empty:
            continue
        out_path = logs_dir / f"{scenario.scenario_id}.parquet"
        df.to_parquet(out_path, index=False)
        extracted += 1
        vlog(f"  {scenario.scenario_id} -> {out_path}  ({len(df)} events)")

    log(f"Extracted {extracted} scenario(s) into logs/")


# ---------------------------------------------------------------------------
# Main flow
# ---------------------------------------------------------------------------
def main() -> None:
    global _VERBOSE

    parser = argparse.ArgumentParser(description="Firmware validation: replay, compare, report.")
    parser.add_argument("--settings", default="settings.json", help="Path to settings JSON file")
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose output")
    parser.add_argument("--report-only", action="store_true", help="Skip replay, just run analysis + report")
    parser.add_argument("--archive", action="store_true", help="Archive logs and extract new baseline after report")
    parser.add_argument("--dev", action="store_true", help="Dev mode: only run 3 test devices for fast iteration")
    args = parser.parse_args()

    _VERBOSE = args.verbose

    # Resolve paths relative to this script's directory
    firmware_dir = Path(__file__).resolve().parent
    settings_path = firmware_dir / args.settings
    if not settings_path.exists():
        print(f"ERROR: Settings file not found: {settings_path}", file=sys.stderr)
        sys.exit(1)

    settings = load_settings(settings_path)
    log(f"signal_replay version: {sr.__version__}")
    log(f"Firmware dir:  {firmware_dir}")
    log(f"Settings:      {settings_path}")

    # --- Read catalog ---
    catalog_path = firmware_dir / settings["catalog_file"]
    catalog = read_catalog(catalog_path)
    log(f"Catalog:       {len(catalog)} rows from {catalog_path.name}")

    # --- Load conflict pairs ---
    conflict_pairs_path = firmware_dir / settings["conflict_pairs_file"]
    conflict_pairs: dict = {}
    if conflict_pairs_path.exists():
        with open(conflict_pairs_path, "r") as f:
            conflict_pairs = json.load(f)
        conflict_pairs = {k: [tuple(pair) for pair in v] for k, v in conflict_pairs.items()}
        vlog(f"Loaded conflict pairs for {len(conflict_pairs)} devices")

    # --- Build suite ---
    suite, scenarios, batches, file_map = build_suite(settings, firmware_dir, catalog, conflict_pairs)

    # Save yaml
    yaml_path = firmware_dir / "test_suite.yaml"
    sr.save_to_yaml(suite, str(yaml_path))
    vlog(f"Saved: {yaml_path}")

    # --- Dev mode filter ---
    if args.dev:
        before = len(scenarios)
        scenarios = [s for s in scenarios if s.scenario_id in DEV_DEVICES]
        suite.scenarios = scenarios
        # Rebuild batches with only the dev devices
        normalized_targets = [normalize_target(t) for t in settings["controller_targets"]]
        dev_batches = []
        for i in range(0, len(scenarios), len(normalized_targets)):
            chunk = scenarios[i:i + len(normalized_targets)]
            batch_num = i // len(normalized_targets) + 1
            targets = normalized_targets[:len(chunk)]
            dev_batches.append(sr.TestBatch(
                batch_id=f"batch_{batch_num}",
                assignments={s.scenario_id: t for s, t in zip(chunk, targets)},
                description=f"Batch {batch_num}: {len(chunk)} scenarios",
            ))
        batches = dev_batches
        suite.batches = batches
        log(f"DEV MODE: filtered {before} -> {len(scenarios)} scenarios ({', '.join(s.scenario_id for s in scenarios)})")

    log(f"Scenarios: {len(scenarios)}  |  Batches: {len(batches)}")

    # --- File readiness ---
    logs_dir = firmware_dir / settings["logs_dir"]
    databases_dir = firmware_dir / settings["databases_dir"]
    ready = sum(1 for r in catalog if file_map[r["TSSU"]]["log"])
    log(f"Logs ready: {ready}/{len(catalog)}")

    if _VERBOSE:
        for r in catalog:
            tssu = r["TSSU"]
            lf = file_map[tssu]["log"]
            df = file_map[tssu]["db"]
            status = "OK" if lf else "--"
            log(f"  {status:>2s}  {tssu:>8s}  log={lf.name if lf else 'MISSING':<30s}  db={df.name if df else '--'}")

    # ======================================================================
    # REPLAY PHASE
    # ======================================================================
    if not args.report_only:
        while True:
            current_batch, completed = detect_current_batch(suite, batches)
            if current_batch is None:
                log("\nAll batches complete! Proceeding to analysis.")
                break

            log(f"\n{'='*70}")
            log(f"BATCH: {current_batch.batch_id}  ({len(current_batch.assignments)} scenarios)")
            log(f"Completed so far: {sorted(completed) if completed else '(none)'}")
            log(f"{'='*70}")
            log("\nLoad these databases onto the controllers:")
            for sid, tgt in current_batch.assignments.items():
                s = next(s for s in scenarios if s.scenario_id == sid)
                db_name = Path(s.database_name).name
                extras = []
                if s.test_type == sr.TestType.CONFLICT:
                    extras.append("CONFLICT" + (" pairs=configured" if s.incompatible_pairs else " (no pairs)"))
                if s.cycle_length:
                    extras.append(f"CL={s.cycle_length}")
                if s.cycle_offset:
                    extras.append(f"Off={s.cycle_offset}")
                if not s.tod_align:
                    extras.append("tod_align=OFF")
                extra_str = "  " + ", ".join(extras) if extras else ""
                log(f"  {tgt:<22s}  <-  {db_name:<20s}{extra_str}")

            log("")
            input("Press ENTER when databases are loaded and controllers are ready...")

            # Controller check
            log("\nChecking controllers...")
            db_lookup = {s.scenario_id: Path(s.database_name).name for s in scenarios}
            b_targets = list(current_batch.assignments.values())
            b_labels = [f"{tgt} / {db_lookup[sid]}" for sid, tgt in current_batch.assignments.items()]

            try:
                wait_for_controllers(b_targets, b_labels)
            except KeyboardInterrupt:
                log("\nController check interrupted. Continuing anyway...")

            # Run replay
            log(f"\nRunning replay for {current_batch.batch_id}...")
            checkpoint = run_batch(suite, current_batch)
            completed_batches = checkpoint.get("completed_batches", [])
            remaining = [b.batch_id for b in batches if b.batch_id not in completed_batches]

            if remaining:
                log(f"\nDone with {current_batch.batch_id}. "
                    f"Completed: {len(completed_batches)}/{len(batches)} batches.")
                log(f"Remaining batches: {remaining}")
                log("Run this script again for the next batch.")
                sys.exit(0)
            else:
                log(f"\nAll {len(batches)} batches complete!")

    # ======================================================================
    # ANALYSIS PHASE
    # ======================================================================
    log(f"\n{'='*70}")
    log("ANALYSIS: Comparing collected output to original logs")
    log(f"{'='*70}")

    results = run_analysis(suite, settings, firmware_dir)
    passed = sum(1 for r in results if r.passed)
    log(f"\nResults: {passed}/{len(results)} passed")

    # ======================================================================
    # REPORT
    # ======================================================================
    log(f"\n{'='*70}")
    log("REPORT: Generating HTML report")
    log(f"{'='*70}")

    report_path = build_report(results, suite)
    log(f"Report saved to: {report_path}")
    log(f"Total images embedded: {sum(len(r.plot_paths) for r in results)}")

    # ======================================================================
    # ARCHIVE (optional)
    # ======================================================================
    if args.archive:
        log(f"\n{'='*70}")
        log("ARCHIVE: Extracting new baseline logs")
        log(f"{'='*70}")
        archive_and_extract(suite, firmware_dir)

    log("\nDone.")


if __name__ == "__main__":
    main()
